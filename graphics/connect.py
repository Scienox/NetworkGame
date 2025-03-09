from ip_untils.ip import IP
from ip_untils.vlsm import VLSM
from ip_untils.helpfunction import cidrRequiermentForHost, cidrRequiermentForSubnetMask
from PySide6.QtWidgets import QTableWidgetItem, QHeaderView, QGridLayout, QSpacerItem, QFileDialog, QMessageBox
from PySide6.QtCore import Qt
from openpyxl import Workbook, load_workbook
from .errorMessage import *


class __Error:
    def __init__(self):
        self._status = False

    def __bool__(self):
        return self.status
    
    def __str__(self):
        return f"{self.status}"

    @property
    def true(self):
        self._status = True

    def false(self):
        self._status = False

    @property
    def status(self):
        return self._status


class __ActionMessage(QMessageBox):
    def __init__(self, window, message):
        super().__init__(window)
        self.setText(message)
        self.yes = self.addButton("Oui", QMessageBox.ActionRole)
        self.no = self.addButton("Non", QMessageBox.ActionRole)

        self.exec()

    def __bool__(self):
        if self.clickedButton() == self.yes:
            return True
        elif self.clickedButton() == self.no:
            return False


def errMessage(window, text):
    dialog = QMessageBox(window)
    dialog.setText(text)
    dialog.addButton("Ok", QMessageBox.ActionRole)
    dialog.exec()


def __selectChoiceCidr(choiceDelimiter, delimiterNetwork):
    cidr = None
    try:
        if choiceDelimiter == 0:
            cidr = int(delimiterNetwork)
        elif choiceDelimiter == 1:
            cidr = cidrRequiermentForSubnetMask(delimiterNetwork)
        elif choiceDelimiter == 2:
            delimiterTmp = int(delimiterNetwork)
            cidr = cidrRequiermentForHost(delimiterTmp)
        return cidr
    except Exception as e:
        raise ValueError(e)


def showIpConfig(ip, choiceDelimiter, delimiterNetwork, table):
    window = table.window()
    error = __Error()
    try:
        cidr = __selectChoiceCidr(choiceDelimiter, delimiterNetwork)
    except Exception as e:
        error.true
        errMessage(window, errSelectCidr(str(e)))
    if not error:
        try:
            IpTmp = IP(ip, cidr)

            methodes = [
                "type", "classIP", "reservation",
                "network", "subMask", "cidr",
                "ipHost", "firstHost", "lastHost",
                "broadcast", "totalHost"
                ]
            for row in range(0, len(methodes)):
                methode = methodes[row]
                targetMethode = getattr(IpTmp, methode)
                element = QTableWidgetItem(str(targetMethode))
                table.setItem(row, 0, element)
        except Exception as e:
            errMessage(table.window(), errIp(str(e)))


def detectHostPart(adressBinary, cidr):
    netPart = ""
    hostPart = ""
    adressBinaryToStr = str("".join(str(octet[bit]) for octet in adressBinary for bit in range(8)))
    dotIndex = 0
    for index in range(32):
        dotIndex += 1
        element = str(adressBinaryToStr[index])
        if index < cidr:
            netPart += element + (" ." if (dotIndex % 8 == 0 and dotIndex < 25) else '') + ' '
        else:
            hostPart += element + (" ." if (dotIndex % 8 == 0 and dotIndex < 25) else '') + ' '
    return netPart, hostPart


def showBinaryInfo(ip, choiceDelimiter, delimiterNetwork, table):
    window = table.window()
    error = __Error()
    try:
        cidr = __selectChoiceCidr(choiceDelimiter, delimiterNetwork)
    except Exception as e:
        error.true
        errMessage(window, errSelectCidr(str(e)))

    if not error:
        try:
            IpTmp = IP(ip, cidr)

            rowsMethodes = [
                "subMask", "network", "ipHost",
                "firstHost", "lastHost", "broadcast"
            ]
            for row in range(0, len(rowsMethodes)):
                methode = rowsMethodes[row]
                targetMethode = getattr(IpTmp, methode)
                targetMethodeB = getattr(IpTmp, methode + "Binary")
                element = QTableWidgetItem(str(targetMethode))
                table.setItem(row, 0, element)
                netPart, hostPart = detectHostPart(targetMethodeB, cidr)
                elementNetPart = QTableWidgetItem(netPart)
                elementHostPart = QTableWidgetItem(hostPart)
                table.setItem(row, 1, elementNetPart)
                table.setItem(row, 2, elementHostPart)
        except Exception as e:
            errMessage(window, errIp(str(e)))


def addNetwork(table):
    columnCount = table.columnCount()
    table.insertColumn(columnCount)
    table.setColumnWidth(columnCount, 100)


def removeNetwork(table):
    columnCurrent = table.currentColumn()
    table.removeColumn(columnCurrent)
    

def readTable(table):
    rows = table.rowCount()
    columns = table.columnCount()
    for row in range(rows):
        line = list()
        for column in range(columns):
            item = table.item(row, column)
            if item:
                line.append(item.text())
            else:
                line.append('')
        yield line


def getNameNetwork(table):
    columns = table.columnCount()
    if columns > 0:
        for column in range(columns):
            name, numberHost = table.item(0, column), table.item(1, column)
            yield name, numberHost
    else:
        raise ValueError("Network not created")


def readVlsm(vlsm):
    for subNetwork in vlsm.subNetworks:
        yield subNetwork


def makeVlsm(tableNetwork, tableVlsm, subDivisedNetwork, choiceDelimiter, delimiterNetwork):
    window = tableNetwork.window()
    error = __Error()
    tableVlsm.setRowCount(0)
    subNetwork = ","
    try:
        subNetwork = subNetwork.join(f"{name.text() if name != None else ''}:{numberHost.text()}" for name,
                                    numberHost in getNameNetwork(tableNetwork))
    except Exception as e:
        error.true
        errMessage(window, errSubnet(str(e)))
    if not error:
        try:
            cidr = __selectChoiceCidr(choiceDelimiter, delimiterNetwork)
        except Exception as e:
            error.true
            errMessage(window, errSelectCidr(str(e)))
    if not error:
        try:
            vlsm = VLSM(subDivisedNetwork, cidr, subNetwork)
            rowCurrent = tableVlsm.rowCount()
            methodes = ["name", "subMask", "network", "totalHost", "cidr"]

            for subnet_work in readVlsm(vlsm):
                tableVlsm.insertRow(rowCurrent)
                updateRowSizeTable(tableVlsm)
                for methodeCurrent in range(len(methodes)):
                    methode = methodes[methodeCurrent]
                    ipInfo = getattr(subnet_work, methode)
                    newItem = QTableWidgetItem(str(ipInfo))
                    tableVlsm.setItem(rowCurrent, methodeCurrent, newItem)
                rowCurrent += 1
        except Exception as e:
            errMessage(window, errVlsm(str(e)))
    updateRowSizeTable(tableVlsm)


def findLayout(widget, layoutType):
    current = widget
    while current is not None:
        layout = current.layout()
        if isinstance(layout, layoutType):
            return layout
        current = current.parentWidget()
    return None


def readGridLayout(layout):
    for row in range(layout.rowCount()):
        for column in range(layout.columnCount()):
            yield layout.itemAtPosition(row, column)


def readRowLayout(layout):
    for row in range(layout.rowCount()):
        yield layout.itemAtPosition(row, 0)


def heightWithoutTarget(target, targetType):
    height = 0
    layout = findLayout(target, targetType)
    for item in readRowLayout(layout):
        if isinstance(item, targetType):
            height += heightWithoutTarget(item, targetType) + layout.spacing()
        elif item is not None:
            itemWidget = item.widget()
            if itemWidget is not None and itemWidget != target:
                height += itemWidget.height() + layout.spacing()
            elif isinstance(item, QSpacerItem):
                height += item.sizeHint().height() + layout.spacing()
    return height


def updateRowSizeTable(table):
    horizontalHeader = table.horizontalHeader()
    screenSize = table.window().height()
    overSize = (20 if not horizontalHeader.isHidden() else 0) + (table.horizontalScrollBar().height() if table.horizontalScrollBarPolicy() == Qt.ScrollBarAlwaysOn else 0)
    totalHeight = (table.rowHeight(0) * table.rowCount()) + overSize
    avaibleHeight = screenSize - heightWithoutTarget(table, QGridLayout) - overSize
    if avaibleHeight <= totalHeight:
        table.setFixedHeight(avaibleHeight)
    else:
        table.setFixedHeight(min(totalHeight, avaibleHeight))


def tableNoResizeRow(table):
    verticalHeader = table.verticalHeader()
    verticalHeader.setSectionResizeMode(QHeaderView.Fixed)


def addAfterRowAddressingPlan(table):
    row = table.rowCount()
    table.insertRow(row)
    updateRowSizeTable(table)


def addBeforeRowAddressingPlan(table):
    table.insertRow(0)
    updateRowSizeTable(table)


def addAfterTRowAddressingPlan(table):
    currentRow = table.currentRow()
    table.insertRow(currentRow + 1 if 0 < currentRow else 1)
    updateRowSizeTable(table)


def addBeforeTRowAddrerssingPLan(table):
    currentRow = table.currentRow()
    table.insertRow(currentRow)
    updateRowSizeTable(table)


def removeRowSelected(table):
    currentRow = table.currentRow()
    table.removeRow(currentRow)
    updateRowSizeTable(table)


def importVlsm(table, tableVlsm):
    window = table.window()
    response = __ActionMessage(window, "Voulez-vous écraser votre plan d'adressage?")
    if response:
        table.setRowCount(0)
        rowCurrent = 0
        for row in range(tableVlsm.rowCount()):
            key = {f"{tableVlsm.horizontalHeaderItem(subNetwork).text()}": f"{tableVlsm.item(row, subNetwork).text()}" for subNetwork in range(tableVlsm.columnCount())}
            name, subnetMask, network = key['Nom'], key['Masque de sous réseau'], key['@Réseau']
            table.insertRow(rowCurrent)
            newItemName = QTableWidgetItem(name)
            newItemSubnetMask = QTableWidgetItem(subnetMask)
            newItemNetwork = QTableWidgetItem(network)

            table.setItem(rowCurrent, 1, newItemName)
            table.setItem(rowCurrent, 4, newItemSubnetMask)
            table.setItem(rowCurrent, 5, newItemNetwork)

            rowCurrent += 1
        updateRowSizeTable(table)


def readFile(_file):
    with open(_file, "r", -1, encoding="utf-8") as file:
        for line in file.readlines():
            yield line.strip().split(',')
    file.close()


def getImportDialog(window, caption, filter):
    fileDialog = QFileDialog(window)
    file, _ = fileDialog.getOpenFileName(window, caption, "", filter)
    return file


def importCSV(table):
    window = table.window()
    response = __ActionMessage(window, "Voulez-vous écraser votre plan d'adressage?")
    if response:
        file = getImportDialog(window, "Sélectionner un fichier", "Fichier CSV (*csv)")
        if file:
            table.setRowCount(0)
            lines = [line for line in readFile(file)]
            lines.pop(0)
            drawAdressingPlan(table, lines)


def readExcel(file):
    workbook = load_workbook(file)

    sheet = workbook.active
    lines = list()
    for row in range(1, sheet.max_row+1):
        element = sheet.cell(row, 1).value
        if (element is not None):
            line = list()
            for column in range(9*2):
                if column % 2 == 0:
                    cell = sheet.cell(row, column+1).value
                    line.append(str(cell).strip() if cell is not None else '')
            lines.append(line)
    return lines


def importExcel(table):
    window = table.window()
    response = __ActionMessage(window, "Voulez-vous écraser votre plan d'adressage?")
    if response:
        file = getImportDialog(window, "Sélectionner un fichier", "Fichier Excel (*xlsx)")
        if file:
            table.setRowCount(0)
            lines = readExcel(file)
            lines.pop(0)
            drawAdressingPlan(table, lines)


def drawAdressingPlan(table, lines):
    for row in range(len(lines)):
        table.insertRow(row)
        line = lines[row]
        for column in range(len(line)):
            element = line[column]
            newItem = QTableWidgetItem(element)
            table.setItem(row, column, newItem)
    updateRowSizeTable(table)


def toImport(table, choiceImport, importThis):
    if choiceImport == 0:
        importVlsm(table, importThis)
    elif choiceImport == 1:
        importCSV(table)
    elif choiceImport == 2:
        importExcel(table)


def getSaveDialog(window, caption, filter):
    dialogFile = QFileDialog(window)
    newFile, _ = dialogFile.getSaveFileName(window, caption, "", filter)
    return newFile


def generateCsv(window, text):
    newFile = getSaveDialog(window, "Créer ou écraser un fichier CSV", "Fichier (*csv)")
    if newFile:
        with open(newFile+(".csv" if not newFile.endswith(".csv") else ''), "w", encoding="utf-8") as file:
            file.writelines(text)
        file.close()


def generateExcel(window, lines, columnHeader):
    newFile = getSaveDialog(window, "Créer ou écraser un fichier Excel", "Fichier (*xlsx)")
    if newFile:
        newWorkBook = Workbook()
        sheet = newWorkBook.active
        rowLines = 0
        for row in range(1, (len(lines))*2, +2):
            for column in range(0, len(columnHeader)):
                setLetters = columnHeader[column]
                startCell = f"{setLetters[0]}{row}"
                endCell = f"{setLetters[1]}{row+1}"
                sheet.merge_cells(f"{startCell}:{endCell}")
                sheet[startCell] = lines[rowLines][column]
            rowLines += 1
        newWorkBook.save(newFile+(".xlsx" if not newFile.endswith(".xlsx") else ''))


def exportToCsv(table, horizontalHeader):
    lines = [line for line in readTable(table)]
    lines.insert(0, horizontalHeader)
    text = ""
    for line in lines:
        text += (",".join(element for element in line)) + "\n"
    generateCsv(table.window(), text)


def exportToExcel(table, horizontalHeader):
    lines = [line for line in readTable(table)]
    lines.insert(0, horizontalHeader)
    columnHeader = [("A","B"),("C","D"),("E","F"),("G","H"),("I","J"),("K","L"),("M","N"),("O","P"),("Q","R")]
    generateExcel(table.window(), lines, columnHeader)


def toExport(table, horizontalHeader):
    window = table.window()
    dialog = QMessageBox(window)
    dialog.setText("Exporter en tant que:")
    csvRole = dialog.addButton("CSV", QMessageBox.ActionRole)
    excelRole = dialog.addButton("Excel", QMessageBox.ActionRole)
    dialog.addButton("Annuler", QMessageBox.ActionRole)
    dialog.exec()

    response = dialog.clickedButton()
    if response == csvRole:
        exportToCsv(table, horizontalHeader)
    elif response == excelRole:
        exportToExcel(table, horizontalHeader)
